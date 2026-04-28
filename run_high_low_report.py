import sqlite3
import pandas as pd
from openpyxl import load_workbook
import threading
import time
import sys



def run_high_low_report(file_path, days, start_date, keep_vba):
    print(f"跑{days}日新高新低資料 -> {file_path}")

    # 嘗試開啟檔案
    try:
        wb = load_workbook(file_path, keep_vba=keep_vba)
    except FileNotFoundError:
        print(f"錯誤：找不到檔案 '{file_path}'，請確認檔案是否存在。")
        input("請按 Enter 鍵結束程式...")
        sys.exit(1)
    except PermissionError:
        print(f"錯誤：檔案 '{file_path}' 可能已被其他程式開啟，請先關閉檔案再執行。")
        input("請按 Enter 鍵結束程式...")
        sys.exit(1)

    # 刪除並重建工作表
    if "大盤" in wb.sheetnames:
        del wb["大盤"]
    wb.create_sheet(title="大盤")
    sheetMarket = wb["大盤"]

    # 表頭設定
    headers = [
        "日期", "成交量", "開盤價", "最高價", "最低價", "收盤價",
        f"股價低於{days}日家數", f"股價高於{days}日家數", "家數",
        f"股價低於{days}日家數比例", f"股價高於{days}日家數比例"
    ]
    for col, header in enumerate(headers, start=1):
        sheetMarket.cell(row=1, column=col, value=header)

    wb.save(file_path)

    conn = sqlite3.connect("stock.db")
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA cache_size = -131072;")   # 128MB cache
    conn.execute("PRAGMA temp_store = MEMORY;")
    conn.execute("PRAGMA mmap_size = 536870912;")  # 512MB memory-mapped I/O

    stop_event = threading.Event()

    def spin():
        start = time.time()
        while not stop_event.is_set():
            elapsed = time.time() - start
            msg = f"\r經過時間：{elapsed:.1f} 秒"
            sys.stdout.write(msg)
            sys.stdout.flush()
            time.sleep(0.5)
        elapsed = time.time() - start
        sys.stdout.write(f"\r完成，耗時：{elapsed:.1f} 秒\n")
        sys.stdout.flush()

    try:
        # 先載入大盤資料
        t = threading.Thread(target=spin)
        t.start()
        query1 = f"""
        SELECT date, close FROM stock_daily
        WHERE date >= '{start_date}' AND stock_id = 'TAIEX'
        ORDER BY date;
        """
        df = pd.read_sql_query(query1, conn)
        stop_event.set()
        t.join()
        print("大盤資料載入完成")

        for excel_row, row in enumerate(df.itertuples(index=False), start=2):
            sheetMarket[f"A{excel_row}"] = row.date
            sheetMarket[f"F{excel_row}"] = row.close

        wb.save(file_path)

        print("資料滾算中...")

        stop_event.clear()
        t = threading.Thread(target=spin)
        t.start()

        # 載入全部資料，用 pandas 滾動視窗取代 correlated subquery
        df_all = pd.read_sql_query(
            f"SELECT date, stock_id, close FROM stock_daily"
            f" WHERE date >= date('{start_date}', '-{days} days')"
            f" ORDER BY stock_id, date", conn
        )
        df_all['date'] = pd.to_datetime(df_all['date'], errors='coerce')
        bad_rows = df_all[df_all['date'].isna()]
        if not bad_rows.empty:
            print(f"\n警告：發現 {len(bad_rows)} 筆日期格式異常的資料，已略過：")
            print(bad_rows.to_string(index=False))
        df_all = df_all.dropna(subset=['date'])

        df_total = (df_all.groupby('date')['stock_id']
                    .nunique().reset_index(name='num_traded_stocks'))

        # rolling(f'{days}D') 使用 calendar days，與原 SQL BETWEEN date(a.date, '-{days-1} days') AND a.date 邏輯相同
        df_idx = df_all.set_index('date')
        df_idx['min_close'] = (df_idx.groupby('stock_id')['close']
                               .transform(lambda x: x.rolling(f'{days}D', min_periods=1).min()))
        df_idx['max_close'] = (df_idx.groupby('stock_id')['close']
                               .transform(lambda x: x.rolling(f'{days}D', min_periods=1).max()))
        df_stats = df_idx.reset_index()

        df_lows = (df_stats[df_stats['close'] == df_stats['min_close']]
                   .groupby('date')['stock_id'].nunique().reset_index(name='num_lows'))
        df_highs = (df_stats[df_stats['close'] == df_stats['max_close']]
                    .groupby('date')['stock_id'].nunique().reset_index(name='num_highs'))

        df = (df_total.merge(df_lows, on='date', how='left')
                      .merge(df_highs, on='date', how='left'))
        df[['num_lows', 'num_highs']] = df[['num_lows', 'num_highs']].fillna(0).astype(int)
        df['low_ratio'] = (df['num_lows'] * 100.0 / df['num_traded_stocks']).round(2)
        df['high_ratio'] = (df['num_highs'] * 100.0 / df['num_traded_stocks']).round(2)
        df = (df[df['date'] >= pd.Timestamp(start_date)]
              .sort_values('date').reset_index(drop=True))
        stop_event.set()
        t.join()
        print(f"{days}日新高新低資料載入完成")

        for excel_row, row in enumerate(df.itertuples(index=False), start=2):
            sheetMarket[f"G{excel_row}"] = row.num_lows
            sheetMarket[f"H{excel_row}"] = row.num_highs
            sheetMarket[f"I{excel_row}"] = row.num_traded_stocks
            sheetMarket[f"J{excel_row}"] = row.low_ratio
            sheetMarket[f"K{excel_row}"] = row.high_ratio

        wb.save(file_path)

    except Exception as e:
        stop_event.set()
        print(f"發生錯誤：{e}")
        input("請按 Enter 鍵結束程式...")
    finally:
        conn.close()
